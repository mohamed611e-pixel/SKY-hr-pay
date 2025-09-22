#!/usr/bin/env python3
"""
Test script to debug migration issues
"""
import os
import sys
import openpyxl
from database_utils import db_manager

def test_excel_access():
    """Test basic Excel file access"""
    print("🔍 Testing Excel file access...")

    if not os.path.exists('generated_payslips.xlsx'):
        print("❌ Excel file 'generated_payslips.xlsx' not found in current directory")
        return False

    try:
        wb = openpyxl.load_workbook('generated_payslips.xlsx', read_only=True)
        print(f"✅ Excel file opened successfully")
        print(f"   Sheets: {wb.sheetnames}")

        ws = wb.active
        print(f"   Active sheet: {ws.title}")
        print(f"   Total rows: {ws.max_row}")

        # Get first few headers
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"   Headers: {headers[:5]}...")  # Show first 5 headers

        wb.close()
        return True

    except Exception as e:
        print(f"❌ Error accessing Excel file: {e}")
        return False

def test_database_connection():
    """Test database connection"""
    print("\n🔍 Testing database connection...")

    try:
        # Try to initialize database
        db_manager.initialize_database()
        print("✅ Database initialized successfully")

        # Test connection
        conn = db_manager.get_connection()
        print("✅ Database connection established")
        conn.close()

        return True

    except Exception as e:
        print(f"❌ Database connection failed: {e}")
        return False

def main():
    print("🧪 Migration Test Script")
    print("=" * 40)

    # Test Excel access
    excel_ok = test_excel_access()

    # Test database connection
    db_ok = test_database_connection()

    print("
📊 Test Results:"    print(f"   Excel file access: {'✅ PASS' if excel_ok else '❌ FAIL'}")
    print(f"   Database connection: {'✅ PASS' if db_ok else '❌ FAIL'}")

    if not excel_ok:
        print("\n💡 To fix Excel issues:")
        print("   1. Ensure 'generated_payslips.xlsx' exists in the current directory")
        print("   2. Check if the file is corrupted")
        print("   3. Verify file permissions")

    if not db_ok:
        print("\n💡 To fix database issues:")
        print("   1. Install MySQL server")
        print("   2. Create database 'sky_hr_payslips'")
        print("   3. Update config.py with correct credentials")
        print("   4. Or use SQLite for testing (modify config.py)")

if __name__ == "__main__":
    main()
