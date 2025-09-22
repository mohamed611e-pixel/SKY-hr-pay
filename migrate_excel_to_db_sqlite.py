#!/usr/bin/env python3
"""
Excel to SQLite Database Migration Script for SKY HR Payslip System
"""
import os
import sys
import openpyxl
import logging
from typing import List, Dict, Any, Tuple
from database_utils_sqlite import db_manager
from config_sqlite import TABLE_EMPLOYEES

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelToDatabaseMigrator:
    """Handles migration of Excel data to SQLite database"""

    def __init__(self, excel_file: str = 'generated_payslips.xlsx'):
        self.excel_file = excel_file
        self.headers = []
        self.total_rows = 0
        self.employees_data = []

    def validate_excel_file(self) -> bool:
        """Validate Excel file exists and is readable"""
        if not os.path.exists(self.excel_file):
            logger.error(f"Excel file not found: {self.excel_file}")
            return False

        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            wb.close()
            logger.info(f"✅ Excel file validation passed: {self.excel_file}")
            return True
        except Exception as e:
            logger.error(f"❌ Excel file validation failed: {e}")
            return False

    def analyze_excel_structure(self) -> Dict[str, Any]:
        """Analyze Excel file structure and content"""
        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            ws = wb.active

            # Get headers
            self.headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            self.total_rows = ws.max_row

            # Find employee-related columns
            column_mapping = self._identify_employee_columns()

            analysis = {
                'filename': self.excel_file,
                'sheet_name': ws.title,
                'total_rows': self.total_rows,
                'headers': self.headers,
                'column_mapping': column_mapping,
                'estimated_employees': self.total_rows - 1  # Excluding header
            }

            logger.info("📊 Excel structure analysis:")
            logger.info(f"   File: {analysis['filename']}")
            logger.info(f"   Sheet: {analysis['sheet_name']}")
            logger.info(f"   Total rows: {analysis['total_rows']}")
            logger.info(f"   Headers: {analysis['headers']}")
            logger.info(f"   Column mapping: {analysis['column_mapping']}")

            wb.close()
            return analysis

        except Exception as e:
            logger.error(f"❌ Failed to analyze Excel structure: {e}")
            return {}

    def _identify_employee_columns(self) -> Dict[str, int]:
        """Identify which columns contain employee data"""
        column_mapping = {}

        # Common variations for employee ID
        hrid_patterns = ['emp id', 'employee id', 'hrid', 'hr id', 'employee_id', 'emp_id']
        nid_patterns = ['national id', 'nid', 'national_id', 'id number', 'national_number']
        name_patterns = ['name', 'employee name', 'employee_name', 'full name', 'full_name']

        for i, header in enumerate(self.headers):
            header_lower = str(header).lower().strip()

            # Check for HRID
            if any(pattern in header_lower for pattern in hrid_patterns):
                column_mapping['hrid'] = i
                logger.info(f"   ✅ Found HRID column: '{header}' at index {i}")

            # Check for NID
            elif any(pattern in header_lower for pattern in nid_patterns):
                column_mapping['nid'] = i
                logger.info(f"   ✅ Found NID column: '{header}' at index {i}")

            # Check for Name
            elif any(pattern in header_lower for pattern in name_patterns):
                column_mapping['name'] = i
                logger.info(f"   ✅ Found Name column: '{header}' at index {i}")

        return column_mapping

    def extract_employee_data(self) -> List[Tuple[str, str, str, str, str]]:
        """Extract employee data from Excel file"""
        if not self._validate_column_mapping():
            return []

        try:
            wb = openpyxl.load_workbook(self.excel_file, read_only=True)
            ws = wb.active

            employees = []
            processed_count = 0
            skipped_count = 0

            logger.info("🔄 Starting data extraction...")

            for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
                try:
                    # Extract data based on column mapping
                    hrid = self._get_cell_value(row, 'hrid')
                    nid = self._get_cell_value(row, 'nid')
                    name = self._get_cell_value(row, 'name')

                    # Validate required fields
                    if not hrid or not nid or not name:
                        logger.warning(f"   ⚠️  Skipping row {row_num}: Missing required data")
                        skipped_count += 1
                        continue

                    # Clean and validate data
                    hrid = str(hrid).strip()
                    nid = str(nid).strip()
                    name = str(name).strip()

                    if len(hrid) < 3 or len(nid) < 5 or len(name) < 2:
                        logger.warning(f"   ⚠️  Skipping row {row_num}: Invalid data format")
                        skipped_count += 1
                        continue

                    employees.append((hrid, nid, name, None, None))  # department and position as None for now
                    processed_count += 1

                    if processed_count % 100 == 0:
                        logger.info(f"   📊 Processed {processed_count} employees...")

                except Exception as e:
                    logger.error(f"   ❌ Error processing row {row_num}: {e}")
                    skipped_count += 1
                    continue

            logger.info("✅ Data extraction completed:")
            logger.info(f"   Total employees extracted: {processed_count}")
            logger.info(f"   Skipped rows: {skipped_count}")

            wb.close()
            return employees

        except Exception as e:
            logger.error(f"❌ Failed to extract employee data: {e}")
            return []

    def _validate_column_mapping(self) -> bool:
        """Validate that required columns are found"""
        required_columns = ['hrid', 'nid', 'name']
        missing_columns = [col for col in required_columns if col not in self._identify_employee_columns()]

        if missing_columns:
            logger.error(f"❌ Missing required columns: {missing_columns}")
            logger.error("Required columns: hrid, nid, name")
            return False

        return True

    def _get_cell_value(self, row, column_type: str) -> Any:
        """Get cell value based on column mapping"""
        column_mapping = self._identify_employee_columns()
        if column_type in column_mapping:
            return row[column_mapping[column_type]].value
        return None

    def migrate_to_database(self, batch_size: int = 100) -> Dict[str, Any]:
        """Migrate Excel data to SQLite database"""
        logger.info("🚀 Starting migration to SQLite database...")

        # Extract data
        employees = self.extract_employee_data()
        if not employees:
            return {'success': False, 'error': 'No valid employee data extracted'}

        # Perform bulk insert
        try:
            successful, failed = db_manager.bulk_insert_employees(employees)

            result = {
                'success': True,
                'total_processed': len(employees),
                'successful_inserts': successful,
                'failed_inserts': failed,
                'success_rate': (successful / len(employees)) * 100 if employees else 0,
                'database_file': db_manager.get_connection().execute("PRAGMA database_list").fetchone()['file']
            }

            logger.info("✅ Migration completed successfully:")
            logger.info(f"   Total processed: {result['total_processed']}")
            logger.info(f"   Successful inserts: {result['successful_inserts']}")
            logger.info(f"   Failed inserts: {result['failed_inserts']}")
            logger.info(f"   Success rate: {result['success_rate']".1f"}%")
            logger.info(f"   Database file: {result['database_file']}")

            return result

        except Exception as e:
            logger.error(f"❌ Migration failed: {e}")
            return {'success': False, 'error': str(e)}

def main():
    """Main migration function"""
    print("🔄 SKY HR Payslip System - Excel to SQLite Database Migration")
    print("=" * 65)

    # Initialize migrator
    migrator = ExcelToDatabaseMigrator()

    # Validate Excel file
    if not migrator.validate_excel_file():
        sys.exit(1)

    # Analyze structure
    analysis = migrator.analyze_excel_structure()
    if not analysis:
        sys.exit(1)

    print("\n📋 Migration Summary:")
    print(f"   Source file: {analysis['filename']}")
    print(f"   Estimated employees: {analysis['estimated_employees']}")
    print(f"   Required columns found: {list(analysis['column_mapping'].keys())}")

    # Confirm migration
    confirm = input("\n❓ Proceed with migration? (y/N): ").lower().strip()
    if confirm not in ['y', 'yes']:
        print("❌ Migration cancelled by user")
        sys.exit(0)

    # Perform migration
    result = migrator.migrate_to_database()

    if result['success']:
        print("
🎉 Migration completed successfully!"        print(f"   Success rate: {result['success_rate']".1f"}%")
        print(f"   Records migrated: {result['successful_inserts']}")
        print(f"   Database file: {result['database_file']}")
        print("\n💡 Next steps:")
        print("   1. Update payslip_site.py to use database_utils_sqlite.py")
        print("   2. Test login functionality with the database")
        print("   3. Consider migrating to MySQL for production use")
    else:
        print(f"\n❌ Migration failed: {result['error']}")
        sys.exit(1)

if __name__ == "__main__":
    main()
