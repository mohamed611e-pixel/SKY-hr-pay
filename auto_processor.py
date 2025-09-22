"""
Auto-upload service for SKY HR Payslip System
Handles automatic processing of Excel and PDF files
"""
import os
import time
import logging
import threading
from typing import Dict, Any
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from config import EXCEL_POLL_INTERVAL, PDF_POLL_INTERVAL, BATCH_SIZE
from database_utils import db_manager

logger = logging.getLogger(__name__)

class AutoUploadService:
    """Service for automatic file processing"""

    def __init__(self):
        self.observer = None
        self.excel_processor = None
        self.pdf_processor = None
        self.running = False
        self.status = "stopped"

    def start(self):
        """Start the auto-upload service"""
        if self.running:
            logger.warning("Auto-upload service is already running")
            return

        try:
            self.running = True
            self.status = "starting"

            # Initialize file system watchers
            self.excel_processor = ExcelFileProcessor()
            self.pdf_processor = PDFFileProcessor()

            # Start background threads
            self.excel_thread = threading.Thread(target=self._excel_processing_loop, daemon=True)
            self.pdf_thread = threading.Thread(target=self._pdf_processing_loop, daemon=True)

            self.excel_thread.start()
            self.pdf_thread.start()

            self.status = "running"
            logger.info("Auto-upload service started successfully")

        except Exception as e:
            self.running = False
            self.status = "error"
            logger.error(f"Failed to start auto-upload service: {e}")
            raise

    def stop(self):
        """Stop the auto-upload service"""
        if not self.running:
            return

        self.running = False
        self.status = "stopping"

        if self.observer:
            self.observer.stop()
            self.observer.join()

        logger.info("Auto-upload service stopped")
        self.status = "stopped"

    def get_status(self) -> Dict[str, Any]:
        """Get current service status"""
        return {
            "running": self.running,
            "status": self.status,
            "excel_processor": self.excel_processor.get_status() if self.excel_processor else None,
            "pdf_processor": self.pdf_processor.get_status() if self.pdf_processor else None
        }

    def _excel_processing_loop(self):
        """Background loop for Excel file processing"""
        logger.info("Excel processing loop started")
        while self.running:
            try:
                self.excel_processor.process_files()
                time.sleep(EXCEL_POLL_INTERVAL)
            except Exception as e:
                logger.error(f"Error in Excel processing loop: {e}")
                time.sleep(10)  # Wait before retrying

    def _pdf_processing_loop(self):
        """Background loop for PDF file processing"""
        logger.info("PDF processing loop started")
        while self.running:
            try:
                self.pdf_processor.process_files()
                time.sleep(PDF_POLL_INTERVAL)
            except Exception as e:
                logger.error(f"Error in PDF processing loop: {e}")
                time.sleep(10)  # Wait before retrying

class ExcelFileProcessor:
    """Handles Excel file processing"""

    def __init__(self):
        self.processed_files = set()
        self.status = "idle"

    def get_status(self) -> Dict[str, Any]:
        """Get processor status"""
        return {
            "type": "excel",
            "status": self.status,
            "processed_files": len(self.processed_files)
        }

    def process_files(self):
        """Process Excel files in upload directory"""
        upload_dir = 'uploaded_payslips'
        if not os.path.exists(upload_dir):
            return

        self.status = "processing"
        try:
            for filename in os.listdir(upload_dir):
                if filename.endswith(('.xlsx', '.xlsm')) and filename not in self.processed_files:
                    filepath = os.path.join(upload_dir, filename)
                    if self._process_excel_file(filepath):
                        self.processed_files.add(filename)

            self.status = "idle"
        except Exception as e:
            logger.error(f"Error processing Excel files: {e}")
            self.status = "error"

    def _process_excel_file(self, filepath: str) -> bool:
        """Process a single Excel file"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(filename=filepath, data_only=True)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            # Find employee columns
            hrid_idx = headers.index('Emp ID') if 'Emp ID' in headers else None
            nid_idx = headers.index('National ID') if 'National ID' in headers else None
            name_idx = headers.index('Name') if 'Name' in headers else None

            if None in (hrid_idx, nid_idx, name_idx):
                logger.error(f"Excel file missing required columns: {filepath}")
                return False

            # Process employee data
            employees = []
            for row in ws.iter_rows(min_row=2):
                hrid = str(row[hrid_idx].value).strip() if row[hrid_idx].value else ''
                nid = str(row[nid_idx].value).strip() if row[nid_idx].value else ''
                name = str(row[name_idx].value).strip() if row[name_idx].value else ''

                if hrid and nid and name:
                    employees.append((hrid, nid, name, None, None))

            # Bulk insert to database
            if employees:
                successful, failed = db_manager.bulk_insert_employees(employees)
                logger.info(f"Processed Excel file {filepath}: {successful} employees inserted, {failed} failed")

                # Move processed file
                processed_dir = os.path.join('uploaded_payslips', 'processed')
                os.makedirs(processed_dir, exist_ok=True)
                os.rename(filepath, os.path.join(processed_dir, os.path.basename(filepath)))

                return True
            else:
                logger.warning(f"No valid employees found in {filepath}")
                return False

        except Exception as e:
            logger.error(f"Error processing Excel file {filepath}: {e}")
            return False

class PDFFileProcessor:
    """Handles PDF file processing"""

    def __init__(self):
        self.processed_files = set()
        self.status = "idle"

    def get_status(self) -> Dict[str, Any]:
        """Get processor status"""
        return {
            "type": "pdf",
            "status": self.status,
            "processed_files": len(self.processed_files)
        }

    def process_files(self):
        """Process PDF files in upload directory"""
        upload_dir = 'uploaded_payslips'
        if not os.path.exists(upload_dir):
            return

        self.status = "processing"
        try:
            for filename in os.listdir(upload_dir):
                if filename.endswith('.pdf') and filename not in self.processed_files:
                    filepath = os.path.join(upload_dir, filename)
                    if self._process_pdf_file(filepath):
                        self.processed_files.add(filename)

            self.status = "idle"
        except Exception as e:
            logger.error(f"Error processing PDF files: {e}")
            self.status = "error"

    def _process_pdf_file(self, filepath: str) -> bool:
        """Process a single PDF file"""
        try:
            import re

            filename = os.path.basename(filepath)

            # Validate filename pattern: Payslip_HRID_MONTH.pdf
            match = re.match(r'^Payslip_(.+)_(\d{2})\.pdf$', filename, re.IGNORECASE)
            if not match:
                logger.warning(f"Invalid PDF filename format: {filename}")
                # Move to invalid folder
                self._move_to_invalid(filepath)
                return False

            hrid = match.group(1)
            month = match.group(2)

            # Check if employee exists in database
            employee = db_manager.get_employee(hrid)
            if not employee:
                logger.warning(f"Employee {hrid} not found in database for PDF: {filename}")
                # Move to invalid folder
                self._move_to_invalid(filepath)
                return False

            # Move to processed folder
            processed_dir = os.path.join('uploaded_payslips', 'processed')
            os.makedirs(processed_dir, exist_ok=True)
            os.rename(filepath, os.path.join(processed_dir, filename))

            logger.info(f"Processed PDF file: {filename} for employee {hrid}")
            return True

        except Exception as e:
            logger.error(f"Error processing PDF file {filepath}: {e}")
            return False

    def _move_to_invalid(self, filepath: str):
        """Move file to invalid folder"""
        invalid_dir = os.path.join('uploaded_payslips', 'invalid')
        os.makedirs(invalid_dir, exist_ok=True)
        os.rename(filepath, os.path.join(invalid_dir, os.path.basename(filepath)))

# Global service instance
auto_upload_service = AutoUploadService()
