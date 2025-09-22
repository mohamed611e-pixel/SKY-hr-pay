#!/usr/bin/env python3
"""
Script to examine the structure of generated_payslips.xlsx
"""
import openpyxl
import sys
from typing import List, Any

def examine_excel_structure():
    """Examine the Excel file structure and content"""
    try:
        # Load the Excel file
        wb = openpyxl.load_workbook('generated_payslips.xlsx')
        print(f"📊 Excel file loaded successfully!")
        print(f"📝 Sheets available: {wb.sheetnames}")

        # Get the active sheet
        ws = wb.active
        print(f"📋 Active sheet: {ws.title}")

        # Get headers
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"📋 Headers ({len(headers)} columns):")
        for i, header in enumerate(headers, 1):
            print(f"   {i"2d"}. {header}")

        # Get total rows
        total_rows = ws.max_row
        print(f"📊 Total rows (including header): {total_rows}")

        # Sample data
        print("
📋 Sample data (first 5 rows):"        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6), 1):
            row_data = [cell.value for cell in row]
            print(f"   Row {i}: {row_data}")

        # Data types analysis
        print("
🔍 Data types analysis:"        if total_rows > 1:
            sample_row = next(ws.iter_rows(min_row=2, max_row=2))
            for i, (header, cell) in enumerate(zip(headers, sample_row)):
                print(f"   Column '{header}': {type(cell.value).__name__}")

        # Check for employee-related columns
        employee_columns = ['Emp ID', 'Employee ID', 'HRID', 'National ID', 'NID', 'Name', 'Employee Name']
        found_columns = [col for col in headers if any(emp_col.lower() in str(col).lower() for emp_col in employee_columns)]
        print("
👥 Employee-related columns found:"        for col in found_columns:
            print(f"   - {col}")

        wb.close()
        return headers, total_rows

    except Exception as e:
        print(f"❌ Error examining Excel file: {e}")
        sys.exit(1)

if __name__ == "__main__":
    print("🔍 Examining Excel file structure...")
    examine_excel_structure()
