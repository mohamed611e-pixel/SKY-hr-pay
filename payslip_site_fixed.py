import os
import re
import logging
import atexit
import signal
import sys
from logging.config import dictConfig
from flask import Flask, request, jsonify, send_from_directory, abort
from flask_cors import CORS
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from pythonjsonlogger import jsonlogger
from hashlib import sha256

# Import auto-upload service components - FIXED: Use SQLite version
from config import AUTO_START
from database_utils_sqlite import db_manager
from auto_processor import auto_upload_service

# Configuration
UPLOAD_FOLDER = 'uploaded_payslips'
ALLOWED_EXTENSIONS = {'pdf'}
EXCEL_DATA = {}
ADMIN_USERNAME = os.getenv('ADMIN_USERNAME', 'Ehab')
ADMIN_PASSWORD_HASH = os.getenv('ADMIN_PASSWORD_HASH', sha256('Ehab611_'.encode()).hexdigest())

# Setup logging with JSON format
dictConfig({
    'version': 1,
    'formatters': {
        'json': {
            '()': jsonlogger.JsonFormatter,
            'fmt': '%(asctime)s %(levelname)s %(message)s'
        }
    },
    'handlers': {
        'console': {
            'class': 'logging.StreamHandler',
            'formatter': 'json'
        }
    },
    'root': {
        'level': 'INFO',
        'handlers': ['console']
    }
})

app = Flask(__name__, static_url_path='/static', static_folder='static')
CORS(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

from flask import send_from_directory

@app.route('/')
def serve_index():
    return send_from_directory('.', 'index.html')

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_excel(file_path):
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active
    employees = {}
    # Adjusted columns: Emp ID (HRID), National ID (NID), Name in first row headers
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    app.logger.info(f"Excel headers: {headers}")
    hrid_idx = headers.index('Emp ID') if 'Emp ID' in headers else None
    nid_idx = headers.index('National ID') if 'National ID' in headers else None
    name_idx = headers.index('Name') if 'Name' in headers else None
    if None in (hrid_idx, nid_idx, name_idx):
        app.logger.error("Excel missing required columns")
        return {}
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        hrid = str(row[hrid_idx].value).strip() if row[hrid_idx].value else ''
        nid = str(row[nid_idx].value).strip() if row[nid_idx].value else ''
        name = str(row[name_idx].value).strip() if row[name_idx].value else ''
        app.logger.info(f"Row {i}: HRID={hrid}, NID={nid}, Name={name}")
        if hrid and nid and name:
            employees[hrid] = {'NID': nid, 'Name': name}
    return employees

@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
       return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and file.filename.lower().endswith(('.xlsx', '.xlsm')):
        filename = secure_filename(file.filename)
        filepath = os.path.join('temp', filename)
        os.makedirs('temp', exist_ok=True)
        file.save(filepath)
        global EXCEL_DATA
        EXCEL_DATA = parse_excel(filepath)
        os.remove(filepath)
        app.logger.info(f"Excel uploaded and parsed: {len(EXCEL_DATA)} employees loaded")
        return jsonify({'message': f'Excel uploaded and parsed: {len(EXCEL_DATA)} employees loaded'})
    else:
        return jsonify({'error': 'Invalid file type, only Excel files allowed'}), 400

@app.route('/login_employee', methods=['POST'])
def login_employee():
    data = request.json
    hrid = data.get('hrid')
    nid = data.get('nid')
    if not hrid or not nid:
        return jsonify({'error': 'Missing HRID or NID'}), 400

    # First check in-memory data (for recently uploaded Excel files)
    employee = EXCEL_DATA.get(hrid)
    if employee and employee['NID'] == nid:
        return jsonify({'success': True, 'name': employee['Name']})

    # Then check database - FIXED: Use SQLite database
    db_employee = db_manager.get_employee(hrid)
    if db_employee and db_employee['nid'] == nid:
        # Update in-memory data for faster future access
        EXCEL_DATA[hrid] = {'NID': db_employee['nid'], 'Name': db_employee['name']}
        return jsonify({'success': True, 'name': db_employee['name']})

    return jsonify({'success': False, 'error': 'Invalid HRID or NID'}), 401

@app.route('/login_admin', methods=['POST'])
def login_admin():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    if not username or not password:
        return jsonify({'error': 'Missing username or password'}), 400
    password_hash = sha256(password.encode()).hexdigest()
    if username == ADMIN_USERNAME and password_hash == ADMIN_PASSWORD_HASH:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': 'Invalid username or password'}), 401

@app.route('/upload_payslips', methods=['POST'])
def upload_payslips():
    if 'files' not in request.files:
        return jsonify({'error': 'No files part'}), 400
    files = request.files.getlist('files')
    saved_files = []
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            # Validate filename pattern Payslip_HRID_MONTH.pdf
            if not re.match(r'^Payslip_.+_\d{2}\.pdf$', filename, re.IGNORECASE):
                continue
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            saved_files.append(filename)
    app.logger.info(f"Uploaded payslips: {saved_files}")
    return jsonify({'uploaded': saved_files})

@app.route('/get_payslip/<hrid>/<month>', methods=['GET'])
def get_payslip(hrid, month):
    # Filename pattern Payslip_HRID_MONTH.pdf
    filename_pattern = f"Payslip_{hrid}_{month}.pdf"
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename_pattern)
    if os.path.exists(filepath):
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename_pattern)
    else:
        return jsonify({'error': 'Payslip not found'}), 404

@app.route('/delete_payslips', methods=['POST'])
def delete_payslips():
    # Admin authentication should be done here (omitted for brevity)
    for filename in os.listdir(app.config['UPLOAD_FOLDER']):
        if filename.lower().endswith('.pdf'):
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], filename))
    app.logger.info("All payslips deleted")
    return jsonify({'message': 'All payslips deleted'})

@app.route('/auto_upload_status', methods=['GET'])
def auto_upload_status():
    """Get auto-upload service status"""
    if not AUTO_START:
        return jsonify({
            'enabled': False,
            'message': 'Auto-upload service is disabled'
        })

    try:
        status = auto_upload_service.get_status()
        return jsonify({
            'enabled': True,
            'status': status
        })
    except Exception as e:
        app.logger.error(f"Error getting auto-upload status: {e}")
        return jsonify({
            'enabled': True,
            'error': str(e)
        })

def signal_handler(signum, frame):
    """Handle shutdown signals gracefully"""
    app.logger.info(f"Received signal {signum}. Shutting down gracefully...")
    cleanup()

def cleanup():
    """Cleanup function for graceful shutdown"""
    try:
        # Stop auto-upload service
        if AUTO_START:
            auto_upload_service.stop()

        # Close database connection
        db_manager.disconnect()

        app.logger.info("Cleanup completed successfully")
    except Exception as e:
        app.logger.error(f"Error during cleanup: {e}")

# Register cleanup function
atexit.register(cleanup)

# Register signal handlers
signal.signal(signal.SIGTERM, signal_handler)
signal.signal(signal.SIGINT, signal_handler)

if __name__ == '__main__':
    # Initialize auto-upload service if enabled
    if AUTO_START:
        try:
            app.logger.info("Starting Auto-Upload Service...")
            auto_upload_service.start()
            app.logger.info("Auto-Upload Service started successfully")
        except Exception as e:
            app.logger.error(f"Failed to start Auto-Upload Service: {e}")
            app.logger.warning("Continuing without auto-upload functionality")

    # Load existing employee data from database - FIXED: Use SQLite database
    try:
        app.logger.info("Loading existing employee data from database...")
        EXCEL_DATA.update(db_manager.get_all_employees())
        app.logger.info(f"Loaded {len(EXCEL_DATA)} employees from database")
    except Exception as e:
        app.logger.error(f"Failed to load employee data from database: {e}")

    port = int(os.environ.get('PORT', 5000))
    app.logger.info(f"Starting Flask application on port {port}")
    app.run(host='0.0.0.0', port=port, debug=False)
